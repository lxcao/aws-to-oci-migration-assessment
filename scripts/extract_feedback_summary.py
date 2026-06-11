#!/usr/bin/env python3
"""Extract customer feedback columns from an AWS-to-OCI migration survey workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl


FEEDBACK_KEYS = ["反馈", "feedback", "customer feedback"]
DIMENSION_KEYS = ["调研维度", "dimension"]
QUESTION_KEYS = ["调研问题", "需要了解的内容", "question"]
INPUT_KEYS = ["客户需提供的信息", "customer input"]
AWS_KEYS = ["aws现状", "aws current"]
OCI_KEYS = ["oci目标", "oci target"]
RISK_KEYS = ["风险", "risk"]


def norm(value: object) -> str:
    return str(value).strip() if value is not None else ""


def find_header_row(ws, max_rows: int = 8) -> tuple[int, list[str]]:
    for row_idx in range(1, min(ws.max_row, max_rows) + 1):
        values = [norm(cell.value) for cell in ws[row_idx]]
        joined = " | ".join(values).lower()
        if any(key in joined for key in FEEDBACK_KEYS) and sum(bool(v) for v in values) >= 3:
            return row_idx, values
    for row_idx in range(1, min(ws.max_row, max_rows) + 1):
        values = [norm(cell.value) for cell in ws[row_idx]]
        if sum(bool(v) for v in values) >= 3:
            return row_idx, values
    return 1, [norm(cell.value) for cell in ws[1]]


def find_col(headers: list[str], keys: list[str]) -> int | None:
    lowered = [h.lower() for h in headers]
    for idx, header in enumerate(lowered):
        for key in keys:
            if key.lower() in header:
                return idx
    return None


def summarize_workbook(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {"source": str(path), "sheets": []}

    for ws in wb.worksheets:
        header_row, headers = find_header_row(ws)
        cols = {
            "dimension": find_col(headers, DIMENSION_KEYS),
            "question": find_col(headers, QUESTION_KEYS),
            "input_required": find_col(headers, INPUT_KEYS),
            "aws_current": find_col(headers, AWS_KEYS),
            "oci_target": find_col(headers, OCI_KEYS),
            "feedback": find_col(headers, FEEDBACK_KEYS),
            "risk": find_col(headers, RISK_KEYS),
        }
        rows = []
        missing_feedback = 0
        for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            cells = [norm(v) for v in values]
            if not any(cells):
                continue
            feedback = cells[cols["feedback"]] if cols["feedback"] is not None and cols["feedback"] < len(cells) else ""
            if not feedback:
                missing_feedback += 1
            row = {}
            for key, idx in cols.items():
                row[key] = cells[idx] if idx is not None and idx < len(cells) else ""
            if feedback or row.get("dimension") or row.get("question"):
                rows.append(row)
        result["sheets"].append({
            "name": ws.title,
            "rows": rows,
            "row_count": len(rows),
            "missing_feedback_count": missing_feedback,
        })
    return result


def to_markdown(summary: dict) -> str:
    lines = [
        "# AWS-to-OCI Migration Survey Feedback Summary",
        "",
        f"Source workbook: `{summary['source']}`",
        "",
        "## Sheet Summary",
        "",
        "| Sheet | Rows | Missing Feedback |",
        "|---|---:|---:|",
    ]
    for sheet in summary["sheets"]:
        lines.append(f"| {sheet['name']} | {sheet['row_count']} | {sheet['missing_feedback_count']} |")

    lines.extend(["", "## Feedback Details", ""])
    for sheet in summary["sheets"]:
        lines.append(f"### {sheet['name']}")
        lines.append("")
        lines.append("| Dimension | Question | Feedback | Initial Risk |")
        lines.append("|---|---|---|---|")
        for row in sheet["rows"]:
            feedback = row.get("feedback") or "[Missing / 待补充]"
            risk = row.get("risk") or ""
            lines.append(
                f"| {row.get('dimension','')} | {row.get('question','')} | {feedback} | {risk} |"
            )
        lines.append("")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Completed migration survey workbook")
    parser.add_argument("--out", help="Markdown output path")
    parser.add_argument("--json", dest="json_out", help="JSON output path")
    args = parser.parse_args()

    summary = summarize_workbook(Path(args.workbook).expanduser().resolve())

    if args.json_out:
        Path(args.json_out).write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    markdown = to_markdown(summary)
    if args.out:
        Path(args.out).write_text(markdown, encoding="utf-8")
    else:
        print(markdown)


if __name__ == "__main__":
    main()
